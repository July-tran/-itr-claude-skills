"""
write_summary.py — Reads all output/assessment_*.json files and writes
a single multi-candidate comparison Excel to output/candidates_summary.xlsx.

Usage:  python write_summary.py [--base-dir PATH]
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── Helpers (same palette as write_excel.py) ───────────────────────────────────

C_HEADER  = "1F3864"
C_WHITE   = "FFFFFF"
C_STRONG  = "C6EFCE"
C_MOD     = "FFEB9C"
C_WEAK    = "FFC7CE"
C_GREEN   = "E2EFDA"
C_ORANGE  = "FCE4D6"
C_SECTION = "D6DCE4"
C_RED_TXT = "FF0000"
C_ORG_TXT = "FF9900"
C_GRN_TXT = "00B050"
C_RANK1   = "FFD700"   # gold
C_RANK2   = "C0C0C0"   # silver
C_RANK3   = "CD7F32"   # bronze


def fill(c):  return PatternFill("solid", fgColor=c)
def font(bold=False, size=11, colour="000000", italic=False):
    return Font(bold=bold, size=size, color=colour, italic=italic)
def border():
    s = Side(style="thin", color="B8CCE4")
    return Border(left=s, right=s, top=s, bottom=s)
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def hcell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(C_HEADER)
    c.font = font(bold=True, colour=C_WHITE, size=10)
    c.alignment = align("center", wrap=True)
    c.border = border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def dcell(ws, row, col, value, bg=None, bold=False, wrap=True, halign="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = align(h=halign, wrap=wrap)
    c.border = border()
    if bg:
        c.fill = fill(bg)
    if bold:
        c.font = font(bold=True)
    return c


# ── Load all candidate JSONs ────────────────────────────────────────────────────

def load_candidates(output_dir: Path) -> list[dict]:
    # JSON data files live in output/data/ to keep output/ Excel-only
    data_dir = output_dir / "data"
    candidates = []
    for f in sorted(data_dir.glob("assessment_*.json")):
        try:
            candidates.append(json.loads(f.read_text(encoding="utf-8")))
        except Exception as e:
            print(f"Warning: could not read {f.name}: {e}", file=sys.stderr)
    return candidates


# ── Sheet 1: Ranking Overview ───────────────────────────────────────────────────

def build_ranking(ws, candidates: list[dict]):
    ws.title = "1 - Ranking Overview"

    headers = [
        ("Rank",            5),
        ("Candidate",      22),
        ("Score",           9),
        ("Match Level",    13),
        ("Recommendation", 28),
        ("Candidate Lvl",  14),
        ("Expected Lvl",   14),
        ("Level Fit",      16),
        ("Matching Skills",18),
        ("Missing Skills", 18),
        ("Risk Signals",   14),
        ("Role",           28),
    ]
    for col, (title, width) in enumerate(headers, 1):
        hcell(ws, 1, col, title, width)

    # Sort by score descending
    sorted_cands = sorted(
        candidates,
        key=lambda d: d.get("score", {}).get("total_score", 0),
        reverse=True,
    )

    rank_fills = {1: C_RANK1, 2: C_RANK2, 3: C_RANK3}

    for rank, d in enumerate(sorted_cands, 1):
        r = rank + 1
        cand     = d.get("candidate", {})
        score    = d.get("score", {})
        level    = d.get("level", {})
        analysis = d.get("analysis", {})
        jd       = d.get("jd", {})

        total     = score.get("total_score", 0)
        match_lvl = score.get("match_level", "")
        rec       = score.get("recommendation", "")

        score_bg = C_STRONG if match_lvl == "Strong" else C_MOD if match_lvl == "Moderate" else C_WEAK
        row_bg   = rank_fills.get(rank, "FFFFFF")

        # Risk signal summary
        risks = analysis.get("risk_signals", [])
        high   = sum(1 for r2 in risks if r2.get("severity") == "High")
        medium = sum(1 for r2 in risks if r2.get("severity") == "Medium")
        risk_str = ""
        if high:   risk_str += f"{high}H "
        if medium: risk_str += f"{medium}M"
        risk_str = risk_str.strip() or "None"

        dcell(ws, r, 1,  rank,                                          bg=row_bg, bold=(rank <= 3), halign="center")
        dcell(ws, r, 2,  cand.get("name", ""),                          bg=row_bg, bold=(rank <= 3))
        c = ws.cell(row=r, column=3, value=f"{total:.0f} / 100")
        c.fill = fill(score_bg); c.font = font(bold=True); c.alignment = align("center"); c.border = border()
        dcell(ws, r, 4,  match_lvl,                                     bg=score_bg, halign="center")
        dcell(ws, r, 5,  rec,                                           bg=score_bg)
        dcell(ws, r, 6,  f"{level.get('candidate_level_title','')} ({level.get('candidate_level','')})", halign="center")
        dcell(ws, r, 7,  f"{level.get('expected_level_title','')} ({level.get('expected_level','')})",   halign="center")
        dcell(ws, r, 8,  level.get("level_fit", ""),                    halign="center")
        dcell(ws, r, 9,  len(analysis.get("matching_skills", [])),      halign="center")
        dcell(ws, r, 10, len(analysis.get("missing_skills", [])),       halign="center")

        risk_colour = C_WEAK if high else C_MOD if medium else C_STRONG
        dcell(ws, r, 11, risk_str, bg=risk_colour, halign="center")
        dcell(ws, r, 12, jd.get("role_title", ""))

        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


# ── Sheet 2: Score Breakdown ────────────────────────────────────────────────────

def build_scores(ws, candidates: list[dict]):
    ws.title = "2 - Score Breakdown"

    headers = [
        ("Candidate",      22),
        ("Total",           9),
        ("Match Level",    13),
        ("Skill Match\n/30", 12),
        ("Experience\n/25", 12),
        ("Tech Stack\n/20", 12),
        ("Level Fit\n/15",  12),
        ("Project\n/10",    12),
        ("AI Yrs",           9),
        ("Total Yrs",        9),
    ]
    for col, (title, width) in enumerate(headers, 1):
        hcell(ws, 1, col, title, width)
    ws.row_dimensions[1].height = 30

    sorted_cands = sorted(
        candidates,
        key=lambda d: d.get("score", {}).get("total_score", 0),
        reverse=True,
    )

    for i, d in enumerate(sorted_cands, 2):
        cand  = d.get("candidate", {})
        score = d.get("score", {})
        bd    = score.get("breakdown", {})
        total = score.get("total_score", 0)
        match = score.get("match_level", "")
        bg    = C_STRONG if match == "Strong" else C_MOD if match == "Moderate" else C_WEAK

        def pct_fill(val, mx):
            ratio = val / mx if mx else 0
            if ratio >= 0.85: return C_STRONG
            if ratio >= 0.65: return C_MOD
            return C_WEAK

        dcell(ws, i, 1,  cand.get("name", ""))
        dcell(ws, i, 2,  f"{total:.0f}",                                     bg=bg, bold=True, halign="center")
        dcell(ws, i, 3,  match,                                               bg=bg, halign="center")
        dcell(ws, i, 4,  bd.get("skill_match", 0),                           bg=pct_fill(bd.get("skill_match",0), 30), halign="center")
        dcell(ws, i, 5,  bd.get("experience_relevance", 0),                  bg=pct_fill(bd.get("experience_relevance",0), 25), halign="center")
        dcell(ws, i, 6,  bd.get("tech_stack_alignment", 0),                  bg=pct_fill(bd.get("tech_stack_alignment",0), 20), halign="center")
        dcell(ws, i, 7,  bd.get("level_fit", 0),                             bg=pct_fill(bd.get("level_fit",0), 15), halign="center")
        dcell(ws, i, 8,  bd.get("project_relevance", 0),                     bg=pct_fill(bd.get("project_relevance",0), 10), halign="center")
        dcell(ws, i, 9,  cand.get("ai_years", ""),                           halign="center")
        dcell(ws, i, 10, cand.get("total_years", ""),                         halign="center")

    ws.freeze_panes = "B2"


# ── Sheet 3: Strengths & Concerns ──────────────────────────────────────────────

def build_strengths_concerns(ws, candidates: list[dict]):
    ws.title = "3 - Strengths & Concerns"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 35

    hcell(ws, 1, 1, "Candidate")
    hcell(ws, 1, 2, "Top Strengths (with evidence)")
    hcell(ws, 1, 3, "Top Concerns / Gaps")
    hcell(ws, 1, 4, "Risk Signals")

    sorted_cands = sorted(
        candidates,
        key=lambda d: d.get("score", {}).get("total_score", 0),
        reverse=True,
    )

    r = 2
    for d in sorted_cands:
        cand     = d.get("candidate", {})
        analysis = d.get("analysis", {})
        score    = d.get("score", {})
        match    = score.get("match_level", "")
        bg       = C_STRONG if match == "Strong" else C_MOD if match == "Moderate" else C_WEAK

        strengths = analysis.get("strengths", [])[:3]
        str_text  = "\n".join(
            f"• {s.get('point','')} — {s.get('evidence','')[:80]}"
            for s in strengths
        )

        weaknesses = analysis.get("weaknesses", [])[:3]
        weak_text  = "\n".join(
            f"• {w.get('point','')} — {w.get('impact','')[:60]}"
            for w in weaknesses
        )

        risks = analysis.get("risk_signals", [])
        sev_icon = {"High": "🔴", "Medium": "🟡", "Low": "🟢"}
        risk_text = "\n".join(
            f"{sev_icon.get(r2.get('severity',''), '•')} {r2.get('signal','')[:60]}"
            for r2 in risks
        )

        dcell(ws, r, 1, cand.get("name", ""), bg=bg, bold=True)
        dcell(ws, r, 2, str_text,  bg=C_GREEN,  wrap=True)
        dcell(ws, r, 3, weak_text, bg=C_ORANGE, wrap=True)
        dcell(ws, r, 4, risk_text, wrap=True)

        lines = max(len(strengths), len(weaknesses), len(risks))
        ws.row_dimensions[r].height = max(40, lines * 30)
        r += 1

    ws.freeze_panes = "A2"


# ── Sheet 4: Skills Comparison ─────────────────────────────────────────────────

def build_skills(ws, candidates: list[dict]):
    ws.title = "4 - Skills Comparison"

    # Collect all unique skills across candidates
    all_matching: set[str] = set()
    all_missing:  set[str] = set()
    for d in candidates:
        a = d.get("analysis", {})
        all_matching.update(s.lower() for s in a.get("matching_skills", []))
        all_missing.update(s.lower()  for s in a.get("missing_skills", []))

    # Build skill rows: skills that appear in at least one candidate's matching or missing
    all_skills = sorted(all_matching | all_missing)

    sorted_cands = sorted(
        candidates,
        key=lambda d: d.get("score", {}).get("total_score", 0),
        reverse=True,
    )

    # Header row
    ws.column_dimensions["A"].width = 28
    hcell(ws, 1, 1, "Skill")
    for col, d in enumerate(sorted_cands, 2):
        name = d.get("candidate", {}).get("name", f"Cand {col-1}")
        hcell(ws, 1, col, name, 16)

    for row, skill in enumerate(all_skills, 2):
        dcell(ws, row, 1, skill.title())
        for col, d in enumerate(sorted_cands, 2):
            a = d.get("analysis", {})
            has = skill in [s.lower() for s in a.get("matching_skills", [])]
            lacks = skill in [s.lower() for s in a.get("missing_skills", [])]
            if has:
                c = ws.cell(row=row, column=col, value="✓")
                c.fill = fill(C_STRONG); c.font = font(colour="00B050", bold=True)
                c.alignment = align("center"); c.border = border()
            elif lacks:
                c = ws.cell(row=row, column=col, value="✗")
                c.fill = fill(C_WEAK); c.font = font(colour=C_RED_TXT, bold=True)
                c.alignment = align("center"); c.border = border()
            else:
                dcell(ws, row, col, "—", halign="center")

    ws.freeze_panes = "B2"


# ── Rank-prefix renaming ───────────────────────────────────────────────────────

def rename_by_rank(output_dir: Path, sorted_cands: list[dict]):
    """
    Rename individual assessment Excel files to reflect current ranking.
    e.g. assessment_Nguyen_Minh_Tri.xlsx → 01_assessment_Nguyen_Minh_Tri.xlsx

    Safe to call multiple times: strips any existing rank prefix before renaming.
    """
    import re
    rank_prefix = re.compile(r"^\d+_")
    total = len(sorted_cands)
    pad = len(str(total))  # zero-pad width: 2 for ≤99, 3 for ≤999

    renamed, skipped = [], []

    for rank, d in enumerate(sorted_cands, 1):
        name = d.get("candidate", {}).get("name", "").replace(" ", "_")
        if not name:
            continue

        # Build the base filename (no prefix)
        base = f"assessment_{name}.xlsx"

        # Find the current file — may already have an old rank prefix
        current = None
        for f in output_dir.glob("*.xlsx"):
            if f.name == "candidates_summary.xlsx":
                continue
            stripped = rank_prefix.sub("", f.name)
            if stripped == base:
                current = f
                break

        if current is None:
            skipped.append(name)
            continue

        new_name = f"{rank:0{pad}d}_{base}"
        new_path = output_dir / new_name

        if current.name != new_name:
            current.rename(new_path)
            renamed.append(f"{current.name} → {new_name}")

    if renamed:
        print(f"[rank] Renamed {len(renamed)} file(s):", file=sys.stderr)
        for r in renamed:
            print(f"  {r}", file=sys.stderr)
    if skipped:
        print(f"[rank] Could not find Excel for: {', '.join(skipped)}", file=sys.stderr)


# ── Main ───────────────────────────────────────────────────────────────────────

def _resolve_base_dir() -> Path:
    for i, arg in enumerate(sys.argv[1:], 1):
        if arg == "--base-dir" and i < len(sys.argv):
            return Path(sys.argv[i + 1]).resolve()
        if arg.startswith("--base-dir="):
            return Path(arg.split("=", 1)[1]).resolve()
    return Path.cwd()


def main():
    base_dir   = _resolve_base_dir()
    output_dir = base_dir / "output"

    if not output_dir.exists():
        print(f"ERROR: output/ directory not found in {base_dir}", file=sys.stderr)
        sys.exit(1)

    candidates = load_candidates(output_dir)
    if not candidates:
        print("No assessment JSON files found in output/. Run individual assessments first.", file=sys.stderr)
        sys.exit(1)

    print(f"Building summary for {len(candidates)} candidate(s)...", file=sys.stderr)

    # Sort once — reused by both the summary sheets and file renaming
    sorted_cands = sorted(
        candidates,
        key=lambda d: d.get("score", {}).get("total_score", 0),
        reverse=True,
    )

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_ranking(wb.create_sheet(), candidates)
    build_scores(wb.create_sheet(), candidates)
    build_strengths_concerns(wb.create_sheet(), candidates)
    build_skills(wb.create_sheet(), candidates)

    out = output_dir / "candidates_summary.xlsx"
    wb.save(str(out))
    print(f"Summary: {out}")

    # Rename individual Excel files to reflect current ranking
    rename_by_rank(output_dir, sorted_cands)


if __name__ == "__main__":
    main()
