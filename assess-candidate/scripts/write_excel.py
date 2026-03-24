"""
write_excel.py — Reads assessment_data.json from the working directory
and writes a formatted Excel assessment report to output/.

Usage:  python write_excel.py
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ── Colour palette ─────────────────────────────────────────────────────────────
C_HEADER  = "1F3864"
C_WHITE   = "FFFFFF"
C_STRONG  = "C6EFCE"
C_MOD     = "FFEB9C"
C_WEAK    = "FFC7CE"
C_GREEN   = "E2EFDA"
C_ORANGE  = "FCE4D6"
C_BLUE    = "D9E1F2"
C_SECTION = "D6DCE4"
C_RED_TXT = "FF0000"
C_ORG_TXT = "FF9900"
C_GRN_TXT = "00B050"

CAT_COLOURS = {
    "skill_validation":     "DEEAF1",
    "deep_probe":           "D9E1F2",
    "project_verification": "E2EFDA",
    "risk_investigation":   "FCE4D6",
    "scenario_based":       "EDD9F0",
    "behavioral":           "FFF2CC",
}


def fill(c): return PatternFill("solid", fgColor=c)
def font(bold=False, size=11, colour="000000", italic=False):
    return Font(bold=bold, size=size, color=colour, italic=italic)
def border():
    s = Side(style="thin", color="B8CCE4")
    return Border(left=s, right=s, top=s, bottom=s)
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def hrow(ws, row, cols, widths=None):
    for c, v in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill      = fill(C_HEADER)
        cell.font      = font(bold=True, colour=C_WHITE)
        cell.alignment = align("center", wrap=True)
        cell.border    = border()
    if widths:
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w


def dcell(ws, row, col, value, bg=None, bold=False, wrap=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = align(wrap=wrap)
    cell.border    = border()
    if bg:
        cell.fill = fill(bg)
    if bold:
        cell.font = font(bold=True)
    return cell


# ── Sheet builders ─────────────────────────────────────────────────────────────

def build_summary(ws, d):
    ws.title = "1 - Candidate Summary"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 70

    score      = d.get("score", {})
    total      = score.get("total_score", 0)
    match_lvl  = score.get("match_level", "")
    rec        = score.get("recommendation", "")
    breakdown  = score.get("breakdown", {})
    cand       = d.get("candidate", {})
    jd         = d.get("jd", {})
    level      = d.get("level", {})
    analysis   = d.get("analysis", {})

    score_bg = C_STRONG if match_lvl == "Strong" else C_MOD if match_lvl == "Moderate" else C_WEAK

    rows = [
        ("Candidate Name",       cand.get("name", "")),
        ("Email",                cand.get("email", "")),
        ("Phone",                cand.get("phone", "")),
        ("Location",             cand.get("location", "")),
        ("Role Applied",         jd.get("role_title", "")),
        ("File",                 cand.get("file_name", "")),
        ("", ""),
        ("ASSESSMENT RESULTS",   None),
        ("Total Score",          f"{total} / 100"),
        ("Match Level",          match_lvl),
        ("Recommendation",       rec),
        ("", ""),
        ("LEVEL MAPPING",        None),
        ("Candidate Level",      f"{level.get('candidate_level_title','')} (Level {level.get('candidate_level','')})"),
        ("Expected Level",       f"{level.get('expected_level_title','')} (Level {level.get('expected_level','')})"),
        ("Level Fit",            level.get("level_fit", "")),
        ("", ""),
        ("SCORE BREAKDOWN",      None),
        ("Skill Match",          f"{breakdown.get('skill_match', 0)} / 30"),
        ("Experience",           f"{breakdown.get('experience_relevance', 0)} / 25"),
        ("Tech Stack",           f"{breakdown.get('tech_stack_alignment', 0)} / 20"),
        ("Level Fit Score",      f"{breakdown.get('level_fit', 0)} / 15"),
        ("Project Relevance",    f"{breakdown.get('project_relevance', 0)} / 10"),
        ("", ""),
        ("MATCHING SKILLS",      ", ".join(analysis.get("matching_skills", [])[:20])),
        ("MISSING SKILLS",       ", ".join(analysis.get("missing_skills", [])[:15])),
        ("", ""),
        ("OVERALL ASSESSMENT",   analysis.get("overall_assessment", "")),
    ]

    for r, (label, value) in enumerate(rows, 1):
        if label == "" and not value:
            continue
        if value is None:  # section header
            cell = ws.cell(row=r, column=1, value=label)
            cell.fill = fill(C_SECTION)
            cell.font = font(bold=True, size=12)
            cell.alignment = align()
            ws.merge_cells(f"A{r}:B{r}")
            continue

        lc = ws.cell(row=r, column=1, value=label)
        lc.font = font(bold=True); lc.alignment = align(); lc.border = border()

        vc = ws.cell(row=r, column=2, value=value)
        vc.alignment = align(wrap=True); vc.border = border()

        if label in ("Total Score", "Match Level", "Recommendation"):
            vc.fill = fill(score_bg)
            vc.font = font(bold=True, size=13 if label == "Total Score" else 11)
        elif label == "MATCHING SKILLS":
            vc.fill = fill(C_GREEN)
        elif label == "MISSING SKILLS":
            vc.fill = fill(C_ORANGE)

    ws.freeze_panes = "A2"


def build_analysis(ws, d):
    ws.title = "2 - Analysis"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 42

    analysis = d.get("analysis", {})
    r = 1

    hrow(ws, r, ["#", "Strength", "Evidence"]); r += 1
    for i, s in enumerate(analysis.get("strengths", []), 1):
        bg = C_GREEN if i % 2 == 0 else "FFFFFF"
        dcell(ws, r, 1, i, bg); dcell(ws, r, 2, s.get("point",""), bg)
        dcell(ws, r, 3, s.get("evidence",""), bg); r += 1

    r += 1
    hrow(ws, r, ["#", "Weakness / Gap", "Impact on Role"]); r += 1
    for i, w in enumerate(analysis.get("weaknesses", []), 1):
        bg = C_ORANGE if i % 2 == 0 else "FFFFFF"
        dcell(ws, r, 1, i, bg); dcell(ws, r, 2, w.get("point",""), bg)
        dcell(ws, r, 3, w.get("impact",""), bg); r += 1

    r += 1
    hrow(ws, r, ["Severity", "Risk Signal", "Explanation"]); r += 1
    sev_colour = {"High": C_RED_TXT, "Medium": C_ORG_TXT, "Low": C_GRN_TXT}
    for risk in analysis.get("risk_signals", []):
        sc = sev_colour.get(risk.get("severity",""), "000000")
        cell = ws.cell(row=r, column=1, value=risk.get("severity",""))
        cell.font = font(bold=True, colour=sc); cell.alignment = align("center")
        cell.border = border()
        dcell(ws, r, 2, risk.get("signal","")); dcell(ws, r, 3, risk.get("explanation","")); r += 1

    r += 1
    insights = d.get("persona_insights", {})
    for persona_name, insight in insights.items():
        hdr = ws.cell(row=r, column=1, value=f"Persona Insights — {persona_name}")
        hdr.fill = fill(C_SECTION); hdr.font = font(bold=True, size=12)
        ws.merge_cells(f"A{r}:C{r}"); r += 1
        if isinstance(insight, dict):
            for key, val in insight.items():
                if isinstance(val, list):
                    val = "\n• ".join([""] + val)
                lc = ws.cell(row=r, column=1, value=key.replace("_"," ").title())
                lc.font = font(bold=True); lc.border = border(); lc.alignment = align()
                vc = ws.cell(row=r, column=2, value=str(val))
                vc.alignment = align(wrap=True); vc.border = border()
                ws.merge_cells(f"B{r}:C{r}"); r += 1

    ws.freeze_panes = "A2"


def build_questions(ws, d):
    ws.title = "3 - Interview Questions"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 65
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 45

    hrow(ws, 1, ["#","Category","Persona Tags","Question","Rationale","Follow-ups"])

    for i, q in enumerate(d.get("questions", []), 2):
        bg = CAT_COLOURS.get(q.get("category",""), "F2F2F2")
        dcell(ws, i, 1, i-1, bg)
        dcell(ws, i, 2, q.get("category","").replace("_"," ").title(), bg)
        dcell(ws, i, 3, ", ".join(q.get("persona_tags",[])), bg)
        dcell(ws, i, 4, q.get("question",""), bg)
        dcell(ws, i, 5, q.get("rationale",""), bg)
        dcell(ws, i, 6, "\n".join(f"• {fu}" for fu in q.get("follow_ups",[])), bg)
        ws.row_dimensions[i].height = max(40, min(len(q.get("question",""))//2, 120))

    ws.freeze_panes = "A2"


# ── Main ───────────────────────────────────────────────────────────────────────

def _resolve_base_dir() -> Path:
    for i, arg in enumerate(sys.argv[1:], 1):
        if arg == "--base-dir" and i < len(sys.argv):
            return Path(sys.argv[i + 1]).resolve()
        if arg.startswith("--base-dir="):
            return Path(arg.split("=", 1)[1]).resolve()
    return Path.cwd()


def main():
    base_dir  = _resolve_base_dir()
    data_file = base_dir / "assessment_data.json"
    if not data_file.exists():
        print(f"ERROR: assessment_data.json not found in {base_dir}", file=sys.stderr)
        sys.exit(1)

    with open(data_file, encoding="utf-8") as f:
        d = json.load(f)

    output_dir = base_dir / "output"
    output_dir.mkdir(exist_ok=True)

    name    = d.get("candidate", {}).get("name", "Unknown").replace(" ", "_")
    xlsx    = output_dir / f"assessment_{name}.xlsx"

    # JSON data stored in output/data/ (not output/) to keep output Excel-only
    data_dir = output_dir / "data"
    data_dir.mkdir(exist_ok=True)
    json_out = data_dir / f"assessment_{name}.json"

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_summary(wb.create_sheet(), d)
    build_analysis(wb.create_sheet(), d)
    build_questions(wb.create_sheet(), d)

    wb.save(str(xlsx))
    with open(json_out, "w", encoding="utf-8") as f:
        json.dump(d, f, indent=2, ensure_ascii=False)

    print(f"Excel: {xlsx}")

    # Update tracking log
    import subprocess
    track_script = Path(__file__).parent / "track.py"
    subprocess.run(
        [sys.executable, str(track_script), "--base-dir", str(base_dir),
         "--add", str(data_file)],
        check=False,
    )


if __name__ == "__main__":
    main()
